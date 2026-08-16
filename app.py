from __future__ import annotations

import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="Pack productivity", page_icon="📦", layout="wide")

RED = "F8696B"
YELLOW = "FFEB84"
GREEN = "63BE7B"
HEADER = "1F4E78"
SUBHEADER = "D9EAF7"


def clean_headers(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    result.columns = (
        result.columns.astype(str)
        .str.replace("\n", " ", regex=False)
        .str.replace("\r", " ", regex=False)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )
    return result


def normalized(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value).lower())


def find_column(columns, candidates: list[str]) -> str | None:
    lookup = {normalized(column): column for column in columns}
    for candidate in candidates:
        key = normalized(candidate)
        if key in lookup:
            return lookup[key]
    return None


def read_matching_sheet(uploaded_file, required_groups: list[list[str]]) -> pd.DataFrame:
    uploaded_file.seek(0)
    workbook = pd.ExcelFile(uploaded_file, engine="openpyxl")
    available = []

    for sheet in workbook.sheet_names:
        preview = clean_headers(pd.read_excel(workbook, sheet_name=sheet, nrows=3))
        available.extend(preview.columns.tolist())
        if all(find_column(preview.columns, group) for group in required_groups):
            return clean_headers(pd.read_excel(workbook, sheet_name=sheet))

    raise ValueError(
        "Не знайдено аркуш із потрібними колонками. Знайдені колонки: "
        + ", ".join(dict.fromkeys(map(str, available)))
    )


def parse_dates(series: pd.Series) -> pd.Series:
    parsed = pd.to_datetime(series, errors="coerce", dayfirst=True)
    if parsed.notna().sum() == 0:
        parsed = pd.to_datetime(series, errors="coerce")
    return parsed


def norm_for_days(days: pd.Series) -> pd.Series:
    return pd.Series(
        pd.cut(
            days,
            bins=[-float("inf"), 6, 13, 20, float("inf")],
            labels=[140, 180, 220, 260],
        ).astype(int),
        index=days.index,
    )


def station_key(value: object) -> tuple:
    text = str(value)
    numbers = tuple(int(number) for number in re.findall(r"\d+", text))
    return numbers, text


def build_report(pack_file, employees_file) -> tuple[pd.DataFrame, list[str], datetime]:
    pack = read_matching_sheet(
        pack_file,
        [
            ["Operator"],
            ["Packing station", "Packing workstation"],
            ["Operation time", "Operation date"],
        ],
    )
    employees = read_matching_sheet(
        employees_file,
        [["Employee ID", "EmployeeID"], ["Create time", "Creation time"]],
    )

    operator_col = find_column(pack.columns, ["Operator"])
    station_col = find_column(pack.columns, ["Packing station", "Packing workstation"])
    operation_col = find_column(pack.columns, ["Operation time", "Operation date"])
    employee_id_col = find_column(employees.columns, ["Employee ID", "EmployeeID"])
    create_col = find_column(employees.columns, ["Create time", "Creation time"])

    pack = pack[[operator_col, station_col, operation_col]].copy()
    pack.columns = ["Operator raw", "Station", "Operation time"]
    employees = employees[[employee_id_col, create_col]].copy()
    employees.columns = ["Employee ID", "Create time"]

    pack["Employee ID"] = (
        pack["Operator raw"].astype(str).str.extract(r"\((\d+)\)")[0].astype("string")
    )
    pack["Employee"] = (
        pack["Operator raw"].astype(str).str.replace(r"\s*\(\d+\)\s*$", "", regex=True).str.strip()
    )
    pack["Operation time"] = parse_dates(pack["Operation time"])
    pack = pack.dropna(subset=["Employee ID", "Station", "Operation time"])
    if pack.empty:
        raise ValueError("Після обробки не залишилося рядків із працівником, станцією та часом.")

    employees["Employee ID"] = (
        employees["Employee ID"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    )
    employees["Create time"] = parse_dates(employees["Create time"])
    employees = employees.drop_duplicates("Employee ID", keep="last")

    report_date = pack["Operation time"].max().normalize()
    pack["Hour"] = pack["Operation time"].dt.floor("h").dt.strftime("%H:%M")
    pack = pack.merge(employees, on="Employee ID", how="left")
    pack["Days worked"] = (report_date - pack["Create time"].dt.normalize()).dt.days
    pack["Days worked"] = pack["Days worked"].fillna(9999).clip(lower=0)
    pack["Norm"] = norm_for_days(pack["Days worked"])

    grouped = (
        pack.groupby(["Station", "Employee", "Employee ID", "Norm", "Hour"], dropna=False)
        .size()
        .rename("Quantity")
        .reset_index()
    )
    grouped["Percent"] = grouped["Quantity"] / grouped["Norm"]

    quantity = grouped.pivot_table(
        index=["Station", "Employee", "Employee ID", "Norm"],
        columns="Hour",
        values="Quantity",
        aggfunc="sum",
        fill_value=0,
    )
    percent = grouped.pivot_table(
        index=["Station", "Employee", "Employee ID", "Norm"],
        columns="Hour",
        values="Percent",
        aggfunc="sum",
        fill_value=0,
    )

    hours = sorted(grouped["Hour"].unique())
    result = quantity.reset_index()
    result["Total"] = result[hours].sum(axis=1)
    base = result[["Station", "Employee", "Employee ID", "Norm", "Total"]].copy()
    for hour in hours:
        base[f"{hour} Quantity"] = quantity[hour].to_numpy()
        base[f"{hour} Percent"] = percent[hour].to_numpy()

    order = sorted(range(len(base)), key=lambda i: (station_key(base.iloc[i]["Station"]), base.iloc[i]["Employee"]))
    base = base.iloc[order].reset_index(drop=True)
    return base, hours, report_date.to_pydatetime()


def color_for_percent(value: float, yellow_from: float) -> str:
    percentage = value * 100
    if percentage >= 100:
        return GREEN
    if percentage >= yellow_from:
        return YELLOW
    return RED


def to_excel(report: pd.DataFrame, hours: list[str], report_date: datetime, yellow_from: float) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="Report", index=False, startrow=2)
        worksheet = writer.book["Report"]
        worksheet["A1"] = "Pack productivity report"
        worksheet["A2"] = f"Report date: {report_date:%d.%m.%Y}"
        worksheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
        worksheet["A1"].fill = PatternFill("solid", fgColor=HEADER)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(report.columns))

        for cell in worksheet[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=HEADER)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        percent_columns = [i + 1 for i, name in enumerate(report.columns) if name.endswith(" Percent")]
        for row in range(4, worksheet.max_row + 1):
            for column in percent_columns:
                cell = worksheet.cell(row=row, column=column)
                cell.number_format = "0%"
                cell.fill = PatternFill("solid", fgColor=color_for_percent(float(cell.value or 0), yellow_from))
                cell.alignment = Alignment(horizontal="center")

        widths = {1: 18, 2: 28, 3: 14, 4: 10, 5: 12}
        for column in range(1, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(column)].width = widths.get(column, 14)
        worksheet.freeze_panes = "F4"
        worksheet.auto_filter.ref = f"A3:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"

        rules = pd.DataFrame(
            {
                "Days worked": ["0–6", "7–13", "14–20", "21+"],
                "Hourly norm": [140, 180, 220, 260],
            }
        )
        rules.to_excel(writer, sheet_name="Norms", index=False)
        norms_sheet = writer.book["Norms"]
        norms_sheet["D1"] = "Color"
        norms_sheet["E1"] = "Meaning"
        legend = [(RED, f"Below {yellow_from:.0f}%"), (YELLOW, f"{yellow_from:.0f}%–99%"), (GREEN, "100% or more")]
        for row, (color, meaning) in enumerate(legend, start=2):
            norms_sheet.cell(row=row, column=4).fill = PatternFill("solid", fgColor=color)
            norms_sheet.cell(row=row, column=5, value=meaning)

    return output.getvalue()


st.title("📦 Pack productivity")
st.write("Завантажте два Excel-файли. Один рядок у Pack data рахується як один виконаний бокс.")

with st.sidebar:
    st.header("Налаштування")
    yellow_from = st.number_input(
        "Жовтий колір від, %",
        min_value=1,
        max_value=99,
        value=80,
        step=1,
        help="100% і більше завжди позначається зеленим.",
    )
    st.caption("Норми: 140 / 180 / 220 / 260 залежно від стажу.")

left, right = st.columns(2)
with left:
    pack_file = st.file_uploader("1. Pack data", type=["xlsx", "xlsm"])
with right:
    employees_file = st.file_uploader("2. Employees", type=["xlsx", "xlsm"])

if st.button("Створити звіт", type="primary", disabled=not (pack_file and employees_file)):
    try:
        with st.spinner("Обробляю дані…"):
            report, hours, report_date = build_report(pack_file, employees_file)
            excel = to_excel(report, hours, report_date, float(yellow_from))
        st.session_state["report"] = report
        st.session_state["excel"] = excel
        st.session_state["report_date"] = report_date
        st.success(f"Готово: {len(report)} працівників, {len(hours)} часових колонок.")
    except Exception as error:
        st.error(str(error))

if "report" in st.session_state:
    report = st.session_state["report"]
    percent_columns = [column for column in report.columns if column.endswith(" Percent")]

    def highlight(value):
        try:
            color = color_for_percent(float(value), float(yellow_from))
            return f"background-color: #{color}; color: #111111"
        except (TypeError, ValueError):
            return ""

    formatters = {column: "{:.0%}" for column in percent_columns}
    styled = report.style.format(formatters).map(highlight, subset=percent_columns)
    st.dataframe(styled, use_container_width=True, height=520, hide_index=True)
    st.download_button(
        "Завантажити Pack_report.xlsx",
        data=st.session_state["excel"],
        file_name=f"Pack_report_{st.session_state['report_date']:%Y-%m-%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
